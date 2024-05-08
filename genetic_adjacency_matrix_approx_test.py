import sys
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from adjacency_matrix_creator import create_adjacency_matrix
from genetic_with_adjacency_matrix_approx import genetic_adjacency_matrix_approx


sys.setrecursionlimit(30000)
folder_path = './DIMACS benchmarks'

REPEAT_COUNT = 1
results_filename = 'genetic_adjacency_matrix_approx_tests.xlsx'
list_dir = os.listdir(folder_path)

if os.path.isfile(results_filename):
    wb = load_workbook(filename=results_filename)
    sheet = wb.active
else:
    wb = Workbook()
    sheet = wb.active

for index, filename in enumerate(list_dir):
    filepath = os.path.join(folder_path, filename)
    data = create_adjacency_matrix(filepath)
    (num_vertices, _, _) = data
    if num_vertices < 600:
        timeout = 10
    elif num_vertices < 1000:
        timeout = 45
    elif num_vertices < 2000:
        timeout = 80
    else:
        timeout = 100

    global_best_sum = 0
    global_best_pos_timeline_sum = []
    time_iter_sum = []
    global_best_at_sum = 0

    for i in range(REPEAT_COUNT):
        (global_best_pos, global_best_pos_timeline, time_iter, global_best_at) = genetic_adjacency_matrix_approx(data, 100, timeout)
        if i == 0:
            global_best_pos_timeline_sum = global_best_pos_timeline
            time_iter_sum = time_iter
        else:
            global_best_pos_timeline_sum = [x + y for x, y in zip(global_best_pos_timeline_sum, global_best_pos_timeline)]
            time_iter_sum = [x + y for x, y in zip(time_iter_sum, time_iter)]
        global_best_sum += sum(global_best_pos)
        global_best_at_sum += global_best_at

    global_best_avg = global_best_sum/REPEAT_COUNT
    global_best_pos_timeline_avg = [element / REPEAT_COUNT for element in global_best_pos_timeline_sum]
    time_iter_sum_avg = [element / REPEAT_COUNT for element in time_iter_sum]
    avg_iter_time = sum(time_iter_sum_avg) / len(time_iter_sum_avg)
    avg_global_best_at = global_best_at_sum / REPEAT_COUNT

    sheet.append([filename])
    sheet.append(["best found", "avg iteration time", "avg best solution found in sec"])
    sheet.append([global_best_avg, avg_iter_time, avg_global_best_at])
    sheet.append(global_best_pos_timeline_avg)
    sheet.append(time_iter_sum_avg)
    sheet.append([" "])

    print(str(index) + filename + " done!")
    wb.save(results_filename)

